"""
This module is responsible for creating and updating excel reports
"""
import datetime
import os
import sys
from datetime import date, datetime
from string import ascii_uppercase  # For Adjusting column Width
from logger import logger
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string, cell
from openpyxl.styles import PatternFill, Alignment, Font  # Protection


class AxExcel:
    def __init__(self, filelocation, rois):
        try:
            self.now = None
            self.rois = rois
            self.filelocation = "{}".format(filelocation)
            self.wb = Workbook()
            self.open_pyxl_workbook()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            logger.error(
                "Error at excel init function : {}|{}|{}|{}".format(exc_type, fname, exc_tb.tb_lineno, e))

    def open_pyxl_workbook(self):
        try:
            """
              This method invokes the create_excel function if it hasn't already been created.
              param lists: None
              return: None
            """
            self.now = datetime.now()
            curdate = self.now.strftime("%Y%m%d")
            if os.path.isfile(self.filelocation):
                logger.info("Excel file already exist")
            else:
                logger.info("Excel file is not exist and create the file")
                self.create_pyxl_workbook()

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            logger.error(
                "Error at open_pyxl_workbook function : {}|{}|{}|{}".format(exc_type, fname, exc_tb.tb_lineno, e))

    def create_pyxl_workbook(self):
        try:
            """
              This method creates an Excel file if it does not already exist at the specified location.
              param lists: None
              return: None
            """
            self.now = datetime.now()
            curdate = self.now.strftime("%Y-%m-%d")
            self.index_sheet = self.wb.create_sheet(title="Index", index=0)
            getbannerpath = os.path.abspath(".\media\e-con-systems-logo.png")
            econ_banner = openpyxl.drawing.image.Image(getbannerpath)
            econ_banner.height = 150
            econ_banner.width = 570
            start_row = 4
            start_column = 1
            end_row = 11
            end_column = 7
            self.index_sheet.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row,
                                         end_column=end_column)
            for col in range(start_column, end_column + 1):
                self.index_sheet.column_dimensions[
                    get_column_letter(col)].width = 12
            for row in range(start_row, end_row + 1):
                self.index_sheet.row_dimensions[row].height = 17
            cell_width = sum(
                [self.index_sheet.column_dimensions[get_column_letter(col)].width for col in
                 range(start_column, end_column + 1)]) * 7
            cell_height = sum(
                [self.index_sheet.row_dimensions[row].height if self.index_sheet.row_dimensions[row].height else 15 for
                 row in
                 range(start_row, end_row + 1)])
            left_offset = (cell_width - econ_banner.width) / 2
            top_offset = (cell_height - econ_banner.height) / 2
            anchor_col = start_column
            anchor_row = start_row
            anchor_col_offset = left_offset / (
                    self.index_sheet.column_dimensions[get_column_letter(anchor_col)].width * 7)
            anchor_row_offset = top_offset / (
                self.index_sheet.row_dimensions[anchor_row].height if self.index_sheet.row_dimensions[
                    anchor_row].height else 15)

            econ_banner.anchor = f"{get_column_letter(anchor_col)}{anchor_row}"
            self.index_sheet.add_image(econ_banner)

            self.index_sheet['I4'] = "Application Name"
            self.index_sheet['I4'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I5'] = "Application Version"
            self.index_sheet['I5'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I6'] = "Created Date"
            self.index_sheet['I6'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I7'] = "Product Name"
            self.index_sheet['I7'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I8'] = "Firmware Version"
            self.index_sheet['I8'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I9'] = "Image Resolution"
            self.index_sheet['I9'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            self.index_sheet['I10'] = "Setup Type"
            self.index_sheet['I10'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            thick_border = Border(left=Side(style='thick'),
                                  right=Side(style='thick'),
                                  top=Side(style='thick'),
                                  bottom=Side(style='thick'))

            al = Alignment(horizontal='center', vertical='center', shrink_to_fit='True')
            for i in range(9, 11):
                for j in range(4, 11):
                    self.index_sheet.cell(row=j, column=i).alignment = al
                    self.index_sheet.cell(row=j, column=i).border = thick_border
                if i == 9:
                    self.index_sheet.column_dimensions["I"].width = 30
                else:
                    self.index_sheet.column_dimensions["J"].width = 40

            ws1 = self.wb.create_sheet("MTF Validation", index=1)
            freeze = ws1['C3']  # freeze row and column
            ws1.freeze_panes = freeze
            ws1['A1'] = "S.No."
            ws1['A1'].font = Font(color="FF0000")
            ws1['A1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['B1'] = "Cycle Start_Time"
            ws1['B1'].font = Font(color="FF0000")
            ws1['B1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['C1'] = "Cycle End_Time"
            ws1['C1'].font = Font(color="FF0000")
            ws1['C1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['D1'] = "Operator Name"
            ws1['D1'].font = Font(color="FF0000")
            ws1['D1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['E1'] = "Mod_brd_sr.no"
            ws1['E1'].font = Font(color="FF0000")
            ws1['E1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['F1'] = "Base_brd_sr.no"
            ws1['F1'].font = Font(color="FF0000")
            ws1['F1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['G1'] = "Product_sr.no"
            ws1['G1'].font = Font(color="FF0000")
            ws1['G1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            ws1['H1'] = "Before gluing"
            ws1['H1'].font = Font(color="FF0000")
            ws1['H1'].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")

            for x in range(1, 8):
                letter = get_column_letter(x)
                ws1.merge_cells('{}1:{}2'.format(letter, letter))

            find_x = ord("G")
            letters = []
            for i in range(len(self.rois)):  # if with regions len(rois) * 5, else len(rois)
                find_x = find_x + 1
                if find_x > 90:
                    value = find_x - 91
                    letters.append('{}{}'.format(chr(value // 26 + 65), chr(value % 26 + 65)))

                else:
                    letters.append(chr(find_x))  # chr() is used to convert ascii to char value

            for j in range(3):
                roi_column = []
                for label in self.rois.keys():
                    roi_column.append('{} MTF Avg.'.format(label))
                for x, y in zip(letters, roi_column):
                    ws1['{}2'.format(x)] = "{}".format(y)
                    ws1['{}2'.format(x)].font = Font(color="FF0000")
                    ws1['{}2'.format(x)].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")
                    ws1.column_dimensions['{}'.format(x)].width = 20
                    ws1.merge_cells('{}1:{}1'.format(letters[0], letters[-1]))

                current_index = column_index_from_string(letters[-1])
                next_index = current_index + 1
                get_colum_ltr = get_column_letter(next_index)
                if j == 0:
                    words = "AVG MTF.BF Gluing"
                elif j == 1:
                    words = "AVG MTF.AF Gluing"
                elif j == 2:
                    words = "AVG MTF.AF Curing"

                ws1['{}1'.format(get_colum_ltr)] = words
                ws1['{}1'.format(get_colum_ltr)].font = Font(color="FF0000")
                ws1['{}1'.format(get_colum_ltr)].fill = PatternFill(start_color="FFFF00", end_color='FFFF00',
                                                                    fill_type="solid")
                ws1.merge_cells('{}1:{}2'.format(get_colum_ltr, get_colum_ltr))

                letters = []
                for x in range(1, len(self.rois) + 1):  # remove *5 without regions
                    next_index_1 = next_index + x
                    get_colum_ltr = get_column_letter(next_index_1)
                    letters.append(get_colum_ltr)

                if j == 0:
                    words = "After Gluing"
                elif j == 1:
                    words = "After Curing"

                ws1['{}1'.format(letters[0])] = words
                ws1['{}1'.format(letters[0])].font = Font(color="FF0000")
                ws1['{}1'.format(letters[0])].fill = PatternFill(start_color="FFFF00", end_color='FFFF00',
                                                                 fill_type="solid")

            letters = []
            for x in range(1, 6):
                next_index_1 = next_index + x
                get_colum_ltr = get_column_letter(next_index_1)
                letters.append(get_colum_ltr)

            ws1['{}1'.format(letters[0])] = "Stations"
            ws1['{}1'.format(letters[0])].font = Font(color="FF0000")
            ws1['{}1'.format(letters[0])].fill = PatternFill(start_color="FFFF00", end_color='FFFF00',
                                                             fill_type="solid")

            station_column = ["Loading", "Displacement sensor", "Focus", 'Gluing', "Curing"]
            for x, y in zip(letters, station_column):
                ws1['{}2'.format(x)] = "{}".format(y)
                ws1['{}2'.format(x)].font = Font(color="FF0000")
                ws1['{}2'.format(x)].fill = PatternFill(start_color="FFFF00", end_color='FFFF00', fill_type="solid")
                ws1.column_dimensions['{}'.format(x)].width = 20
                ws1.merge_cells('{}1:{}1'.format(letters[0], letters[-1]))

            letters = []
            for x in range(1, 3):
                next_index_2 = next_index_1 + x
                get_colum_ltr = get_column_letter(next_index_2)
                letters.append(get_colum_ltr)

            ws1['{}1'.format(letters[0])] = "Overall Result"
            ws1['{}1'.format(letters[0])].font = Font(color="FF0000")
            ws1['{}1'.format(letters[0])].fill = PatternFill(start_color="FFFF00", end_color='FFFF00',
                                                             fill_type="solid")
            ws1.merge_cells('{}1:{}2'.format(letters[0], letters[0]))

            ws1['{}1'.format(letters[1])] = "Remarks"
            ws1['{}1'.format(letters[1])].font = Font(color="FF0000")
            ws1['{}1'.format(letters[1])].fill = PatternFill(start_color="FFFF00", end_color='FFFF00',
                                                             fill_type="solid")
            ws1.merge_cells('{}1:{}2'.format(letters[1], letters[1]))

            thick_border = Border(left=Side(style='thick'),
                                  right=Side(style='thick'),
                                  top=Side(style='thick'),
                                  bottom=Side(style='thick'))

            al = Alignment(horizontal='center', vertical='center', shrink_to_fit='True')
            for column in range(1, ws1.max_column + 1):
                letter = get_column_letter(column)
                if letter == 'A':
                    ws1.column_dimensions[letter].width = 10
                elif letter == 'B' or letter == 'C' or letter == 'I' or letter == 'J' or letter == 'N':
                    ws1.column_dimensions[letter].width = 30
                else:
                    ws1.column_dimensions[letter].width = 25

            ws1.column_dimensions[letters[1]].width = 100

            if len(self.rois) == 5:  # column = 29(for region average only) column = 89(with each region & avg)
                column = 24  # 95(with regions) #35without regions

            elif len(self.rois) == 7:
                column = 42  # 123(with regions) #42without regions

            elif len(self.rois) == 9:
                column = 47  # 151(with regions) #47without regions

            elif len(self.rois) == 13:
                column = 54  # 207(with regions) #54without regions

            for i in range(1, column + (len(self.rois)) + 4):
                for j in range(1, 3):
                    ws1.cell(row=j, column=i).alignment = al
                    ws1.cell(row=j, column=i).border = thick_border

            logger.info("Saving the file in the current directory")
            ws1.protection.password = "sms"
            ws1.protection.enable()
            self.wb.save('{}'.format(self.filelocation))
            logger.info("File saved successfully")

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            logger.error(
                "Error at create_pyxl_workbook function : {}|{}|{}|{}".format(exc_type, fname, exc_tb.tb_lineno, e))

    def write_data(self, basic_details, mtf_data):
        try:
            """
              This method is used to append the mtf values and index values in excel.
              param lists: None
              return: Bool
            """
            self.now = datetime.now()
            curdate = self.now.strftime("%Y%m%d")
            myworkbook = openpyxl.load_workbook(self.filelocation)

            print(basic_details, mtf_data)

            worksheet_index = myworkbook["Index"]
            worksheet_index['J4'] = "e-ZeeFocus"
            worksheet_index['J4'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J5'] = basic_details[0]
            worksheet_index['J5'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J6'] = basic_details[1]
            worksheet_index['J6'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J7'] = basic_details[2]
            worksheet_index['J7'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J8'] = basic_details[3]
            worksheet_index['J8'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J9'] = basic_details[4]
            worksheet_index['J9'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_index['J10'] = basic_details[5]
            worksheet_index['J10'].fill = PatternFill(start_color="D3D3D3", end_color='D3D3D3', fill_type="solid")

            worksheet_py = myworkbook["MTF Validation"]
            al = Alignment(horizontal='center', vertical='center', shrink_to_fit='True')
            last_row_offset = worksheet_py.max_row + 1

            last_serial_value = worksheet_py.cell(worksheet_py.max_row, 1)
            incremented_value = last_serial_value.value
            incremented_value = str(incremented_value)

            if incremented_value is not None and incremented_value.isnumeric():
                incremented_value = int(incremented_value)
                incremented_value = incremented_value + 1  # Auto increment serial number
                mtf_data.insert(0, incremented_value)

            elif worksheet_py["A1"].value == "S.No.":
                incremented_value = 1
                mtf_data.insert(0, incremented_value)

            else:
                incremented_value += 1
                mtf_data.insert(0, incremented_value)

            print("Final excel result list: {}".format(mtf_data))

            for i in range(0, len(mtf_data)):
                worksheet_py.cell(last_row_offset, i + 1).value = mtf_data[i]
                worksheet_py.cell(last_row_offset, i + 1).alignment = al

            worksheet_index.protection.password = "sms"
            worksheet_index.protection.enable()
            myworkbook.save(self.filelocation)
            return True

        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            logger.error("Error at write_data function : {}|{}|{}|{}".format(exc_type, fname, exc_tb.tb_lineno, e))
            return False

    def isFileOpened(self):
        try:
            """
              This method is used to verify if the excel file is opened or not.
              param lists: None
              return: Bool
            """
            self.now = datetime.now()
            curdate = self.now.strftime("%Y%m%d")
            myworkbook = openpyxl.load_workbook(self.filelocation)
            myworkbook.save(self.filelocation)
            return False

        except PermissionError:
            logger.error("Excel file is opened")
            return True